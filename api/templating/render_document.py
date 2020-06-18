from copy import copy


from docxtpl import DocxTemplate
from jinja2 import Template
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from pptx_template import render


def render_docx(type, data, output_name):
    """
        data = { 'consultant_name' : "Jean" }
        output_name = "rm-jean"
    """
    doc = DocxTemplate("data/templates/{type}.docx".format(type=type))
    doc.render(data)
    output_path = "data/output/" + output_name + ".docx"
    doc.save(output_path)
    return output_path

def render_xlsx(type, data, output_name):
    """
        data = {
            "pay": "180",
            "jeh": "2",
        }
        output_name = 'bv-jean'
    """
    filename = "data/templates/{type}.xlsx".format(type=type)
    sheetname = 'template'

    wb_template = load_workbook(filename=filename)
    ws = wb_template[sheetname]

    wb = Workbook()
    ws_out = wb.active
    ws_out.title = output_name

    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            content = ws[get_column_letter(col) + str(row)].value
            if content:

                # Use french decimal format
                if isinstance(content, float):
                    content = str(content)
                    content = content.replace('.', ',')

                template = Template(str(content))
                rendered = template.render(data)
                ws_out[get_column_letter(col) + str(row)] = rendered
                ws_out[get_column_letter(col) + str(row)].font = copy(ws[get_column_letter(col) + str(row)].font)

    output_path = "data/output/" + output_name + '.xlsx'
    wb.save(filename=output_path)
    return output_path


def render_pptx(type, data, output_name):
    input_path = "data/templates/{type}.pptx".format(type=type)
    output_path = "data/output/" + output_name + ".pptx"
    render.render_pptx(input_path, data, output_path)
    return output_path
